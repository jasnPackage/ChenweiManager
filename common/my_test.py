#coding:utf-8

import xlrd,os,json,ast,re,datetime
import requests
import openpyxl
from openpyxl.styles import Font
# from xlutils.copy import copy
from xlrd import xldate_as_tuple
from config import readConfig
from common.logger import Log
from case.user_login import Login
from openpyxl.styles import Alignment
from common.excel_report import create_excel_report
from common import parameters_handle


'''
获取测试用例data所在的目录
'''
d = os.path.dirname(__file__) #返回当前文件所在目录（common文件夹路径）
parent_path = os.path.dirname(d) #返回common的父级目录
data_path = os.path.join(parent_path,'data') #返回data所在目录
data_path1 = os.listdir(data_path) #返回data目录下所有的文件

s = requests.session()
lon = Login(s)
log = Log()



def api_data():

    pass_case = 0  #通过用例数标识
    fail_case = 0  #失败用例数标识


    # start_time = datetime.now()  #测试用例执行开始时间
    # now_time = start_time.strftime('%Y-%m-%d %H:%M:%S') #开始时间格式转换

    for filename in data_path1:
        book = xlrd.open_workbook(os.path.join(data_path,filename))

        '''
        使用xlwt操作excel，xlwt只支持excel2007以下版本
        '''

        # wb = copy(book)
        # ws = wb.get_sheet(1)

        '''
        使用openpyxl操作excel，openpyxl支持excel2007以上版本
        '''
        wb = openpyxl.load_workbook(os.path.join(data_path,filename))
        ws = wb.worksheets[1]
        font_green = Font(color="37b400")
        font_red = Font(color="ff0000")


        '''
        获取excel文件中接口信息
        '''
        table = book.sheet_by_index(0) #通过索引，获取相应的列表，这里表示获取excel的第一个列表
        inf_name = table.row_values(1)[0] #返回接口名称
        inf_address = table.row_values(1)[1] #返回接口地址
        inf_mode = table.row_values(1)[2] #返回请求方式

        '''
        获取excel文件中测试用例信息，将A-K列转换成字典存储
        '''
        sheet = book.sheet_by_index(1)  # 通过索引，获取相应的列表，这里表示获取excel的第二个列表
        nrows = sheet.nrows  # 获取所有行数
        ncols = sheet.ncols  # 获取所有列数
        key = sheet.row_values(0)

        for i in range(1,nrows):
            value = sheet.row_values(i)
            case_dict = dict(zip(key,value))

            # 读取测试用例中的用例级别
            case_level = case_dict["用例级别"]
            # 读取run_config.ini中设置用例运行级别
            runCase_level = readConfig.runCase_level

            '''
            获取cfg.ini配置文件中接口公共信息（ip和port）和run_config.ini配置文件中
            运行信息
            '''
            runmode = readConfig.runmode
            if runmode == 'test':
                ip = readConfig.test_ip  # 获取配置文件中接口ip
                i_port = readConfig.test_port  # 获取配置文件中接口port
            elif runmode == 'pro':
                ip = readConfig.pro_ip  # 获取配置文件中接口ip
                i_port = readConfig.pro_port  # 获取配置文件中接口port
            else:
                raise ValueError("run_config.ini文件运行环境(runmode)错误")

            url = "http://" + ip + ":" + i_port + inf_address
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0",
                "X-Requested-With": "XMLHttpRequest",
                "Connection": "keep-alive"
            }


            par = case_dict["请求体body"]
            login_userid = case_dict["用户userid"]



            if login_userid == '是':
                new_par = parameters_handle.parameters_handle(par)
            else:
                new_par = par


            # 统计每个测试用例的执行时间
            # 用例执行开始时间
            now_time = datetime.datetime.now()
            start_time = datetime.datetime.strptime(str(now_time),'%Y-%m-%d %H:%M:%S.%f')  # 测试用例执行开始时间

            '''
            判断请求方式是GET还是POST，并且判断测试用例预期结果与实际响应一致，所有接口请求前先调用登录接口
            '''
            if case_level == runCase_level or runCase_level == "ALL":

                if inf_mode == 'GET':
                    r = s.get(url, params=new_par)
                    result = r.text
                elif inf_mode == 'POST':
                    if case_dict["请求头header"] == "form":
                        headers["Content-Type"] = "application/x-www-form-urlencoded"
                    elif case_dict["请求头header"] == "json":
                        headers["Content-Type"] = "application/json"
                    else:
                        raise ValueError("请求头header不能识别")

                    r = s.post(url, data=new_par, headers=headers)
                    result = r.text

                # 先清空实际响应单元格的内容
                ws.cell(row=i + 1, column=ncols - 3, value='')
                # 设置实际响应写入的内容中对齐方式，居中对齐
                align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 纯色填充
                # 将实际响应写入测试用例文件中
                ws.cell(row=i + 1, column=ncols - 3, value=result).alignment = align

                log.info("\n用例文件：%s\n用例编号：%s\n接口名称：%s\n用例名称：%s\n用例描述：%s\n请求内容：%s\n实际响应：%s" % (
                filename,case_dict["用例编号"], inf_name, case_dict["用例名称"], case_dict["用例描述"], new_par, result))

                checkPoints = case_dict["检查点"]
                list_checkPoints = ast.literal_eval(checkPoints)  # string转换成list

                # 设置检查点通过失败次数
                checkPoint_pass = 0
                checkPoint_fail = 0

                for checkPoint in list_checkPoints:
                    # 将dict转换成json，并且去除转换成json后冒号后的空格
                    json_checkPoint = json.dumps(checkPoint, separators=(',', ':'), ensure_ascii=False)
                    # 通过正则取{}中间的值
                    new_checkPoint = (re.findall('{(.*)}', json_checkPoint))[0]
                    if new_checkPoint in result:
                        checkPoint_pass += 1
                        log.info("检查点:%s--pass" % new_checkPoint)
                        log.info("----------------")
                    else:
                        # ws.write(i,ncols-1,'fail'.encode('utf-8'))
                        # ws.cell(row=i + 1, column=ncols, value='Fail').font = font_red
                        checkPoint_fail += 1
                        log.error("检查点:%s--fail" % new_checkPoint)
                        log.info("----------------")

                if checkPoint_pass == len(list_checkPoints):
                    pass_case += 1
                    ws.cell(row=i + 1, column=ncols, value='Pass').font = font_green
                else:
                    fail_case += 1
                    ws.cell(row=i + 1, column=ncols, value='Fail').font = font_red

                # 用例执行结束时间
                now_time = datetime.datetime.now()
                end_time = datetime.datetime.strptime(str(now_time), '%Y-%m-%d %H:%M:%S.%f')
                # 用例执行耗费时间
                elapse_time = str((end_time - start_time)) + "秒"
                # 先清空执行耗时单元格的内容
                ws.cell(row=i + 1, column=ncols - 1, value='')

                # 将用例执行耗时写入测试用例中
                ws.cell(row=i + 1, column=ncols - 1, value=elapse_time).alignment = align


                # 最近一次用例执行时间
                execution_time = now_time.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=i + 1, column=ncols - 2, value=execution_time).alignment = align



                # 保存测试用例文件
                wb.save(os.path.join(data_path, filename))

            else:
                log.info("用例编号:%s--非指定运行级别的用例，不执行"%case_dict["用例编号"])

        total_case = pass_case + fail_case  # 执行用例总数
        log.info("执行测试用例数：%s  通过用例数：%s  失败用例数：%s" % (total_case, pass_case, fail_case))

        # 创建Excel测试报告
        create_excel_report(now_time, pass_case, fail_case, "%.2f%%" % (pass_case / (total_case) * 100))





api_data()